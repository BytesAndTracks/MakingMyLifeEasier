import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def clean_text(value):
    if isinstance(value, str):
        return value.strip().upper()
    return value

def remove_fill_colors(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(fill_type=None)

def process_file(file_path, output_path):
    xls = pd.ExcelFile(file_path)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df.dropna(how='all', inplace=True)
            # Aplicar limpeza em todas as células
            df = df.applymap(clean_text)
            # Corrigir nome das colunas
            df.columns = [col.strip().upper() if isinstance(col, str) else col for col in df.columns]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Remove formatação do arquivo salvo
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        remove_fill_colors(ws)
    wb.save(output_path)

if __name__ == "__main__":
    arquivos = [
        r"C:\Users\Rubens\Downloads\TABELA MASTER- FORNECEDORES  (SENDO ATUALIZADA EM 2025- Eliane.xlsx",
        r"C:\Users\Rubens\Documents\Tremed\Planilhas Tremed\Master_Fornecedor.xlsx",
        r"C:\Users\Rubens\Documents\Tremed\Planilhas Tremed\Tabela_Master.xlsx"
    ]
    for arquivo in arquivos:
        nome_saida = arquivo.replace(".xlsx", "_formatado.xlsx")
        process_file(arquivo, nome_saida)
        print(f"Arquivo processado e salvo como: {nome_saida}")
